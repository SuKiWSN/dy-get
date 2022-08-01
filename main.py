import requests
from bs4 import BeautifulSoup
from explorer import request
from openpyxl import Workbook
import openpyxl
import json


def userDescription(profileUrl):
    res = request(profileUrl).get()
    soup = BeautifulSoup(res.text, 'lxml')
    userId = soup.find('span', {'class': 'kbjj_psh'})
    if userId:
        userId = userId.text
        userId = userId.replace('抖音号：', '')
    else:
        userId = ''
    ipBelong = soup.find('span', {'class': 'WXnH80ht'})
    if ipBelong:
        ipContent = ipBelong.text
        ipBelong = ipContent.replace('IP属地：', '')
    else:
        ipBelong = ''
    return userId, ipBelong

def getVideoId(fpath):
    f = open(fpath, 'r')
    js = json.loads(f.read())
    data = js['data']
    idList = []
    for var in data:
        aweme_info = var['aweme_info']
        aweme_id = aweme_info['aweme_id']
        idList.append(aweme_id)
    return idList


def insert(videoId, i):
    res = request('https://www.douyin.com/video/'+videoId).get()
    soup = BeautifulSoup(res.text, 'lxml')
    divUnit = soup.find_all('div', attrs={'class': 'YzbzCgxU'})

    for var in divUnit:
        userLink = 'https:' + var.find('a')['href']
        userSpan = var.find('span', attrs={'class': 'Nu66P_ba'})
        userNickName = userSpan.text
        createTime = var.find('p', attrs={'class': 'dn67MYhq'}).text
        description = var.find('span', attrs={'class': 'VD5Aa1A1'}).text
        itemTag = var.find('span', attrs={'class': 'yhNyqX95'})
        if itemTag:
            itemAttach = itemTag.text
            if itemAttach == '作者':
                continue
        userId, ipBelong = userDescription(userLink)
        print(userNickName, createTime, description, userId, ipBelong)
        ws['A{}'.format(i)] = userNickName
        ws['B{}'.format(i)] = createTime
        ws['C{}'.format(i)] = description
        ws['D{}'.format(i)] = userId
        ws['E{}'.format(i)] = ipBelong
        ws['F{}'.format(i)] = '=HYPERLINK("{}", "查看主页")'.format(userLink)
        i += 1
    wb.save('user.xlsx')
    return i

if __name__ == '__main__':
    wb = openpyxl.load_workbook('user.xlsx')
    ws = wb.active
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['D'].width = 20
    i = 1
    while ws['A{}'.format(i)].value != None:
        i += 1
    index = ['昵称', '创建日期', '评论', '抖音号', 'ip归属地', '查看主页']
    s = 'A'
    for j in range(len(index)):
        ws['{}1'.format(s)] = index[j]
        s = chr(ord(s) + 1)

    idList = getVideoId('/Users/wanghanyu/Downloads/single')
    for id in idList:
        i = insert(id, i)